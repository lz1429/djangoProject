# -*- coding:utf-8 -*-
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from django import forms

from app01 import models
from utils.pagination import Pagination
from utils.bootstrap import BootStrap


class DepartModelForm(BootStrap, forms.ModelForm):
    """ 部门 ModelForm 类  """

    class Meta:
        model = models.Department
        fields = ['title']
        # 取表中所有的字段（没包括id）：fields = "__all__"
        # 取除了表中的某个字段的所有字段：exclude = ['level']


def depart_list(request):
    """ 部门列表 """
    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["title__contains"] = search_data
    queryset = models.Department.objects.filter(**data_dict)
    page_object = Pagination(request, queryset)
    form = DepartModelForm()
    context = {
        "form": form,
        "search_data": search_data,
        "queryset": page_object.page_queryset,
        "page_string": page_object.html(),
    }
    return render(request, 'depart_list.html', context)


@csrf_exempt
def depart_add(request):
    """ 添加部门 """
    form = DepartModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, 'error': form.errors})


def depart_detail(request):
    """ 编辑部门功能前戏：获取部门详细信息   """
    depart_id = request.GET.get("depart_id")
    row_dict = models.Department.objects.filter(id=depart_id).values("title").first()
    if not row_dict:
        return JsonResponse({"status": False, 'error': "数据不存在。"})
    result = {
        "status": True,
        "data": row_dict
    }
    return JsonResponse(result)


@csrf_exempt
def depart_edit(request):
    """ 编辑部门 """
    depart_id = request.GET.get("depart_id")
    row_object = models.Department.objects.filter(id=depart_id).first()
    if not row_object:
        return JsonResponse({"status": False, 'tips': "数据不存在，请刷新重试。"})
    form = DepartModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": False, 'error': form.errors})


def depart_delete(request):
    """ 删除部门 """
    depart_id = request.GET.get('depart_id')
    exists = models.Department.objects.filter(id=depart_id).exists()
    if not exists:
        return JsonResponse({"status": False, 'error': "删除失败，数据不存在。"})
    models.Department.objects.filter(id=depart_id).delete()
    return JsonResponse({"status": True})


def depart_multipart_add(request):
    """ 批量添加部门（上传Excel文件：第一列为部门名称） """
    file_object = request.FILES.get("excel")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]  # 第一个工作表
    for row in sheet.iter_rows(min_row=1):  # 第一行开始循环
        text = row[0].value  # 每一行的第一列内容
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect('/depart/list/')
